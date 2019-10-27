from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional, Dict, Tuple, Type, List

from openpyxl import Workbook
from gspread_downloader import get_spreadsheet
import app_config as cfg


@dataclass
class DataForPlots:
    """
    Class for storing different lists that represents axes of a plot to
    be rendered and a list with smartphones to be highlighted
    """
    y_axis_names: List[str] = field(default_factory=list)
    x_axis_values: List[int] = field(default_factory=list)
    x_axis_values2: List[Optional[int]] = field(default_factory=list)
    x_axis_values3: List[Optional[int]] = field(default_factory=list)
    highlighted_smartphones: List[int] = field(default_factory=list)

    def all_axes_used(self) -> bool:
        """
        Finds out whether there is any empty list among the axes
        ot not
        :return: true or false respectively
        """
        for key, value in self.__dict__.items():
            if 'axis' in key:
                if not value:
                    return False
        return True

    def get_axes(self) -> Tuple[List[str], List[int], List[Optional[int]],
                                List[Optional[int]]]:
        """
        Get all attributes which represent axes for a plot
        :return: tuple of axes
        """
        return (self.y_axis_names, self.x_axis_values, self.x_axis_values2,
                self.x_axis_values3)


class Benchmark:

    """
    Parent class for any benchmark - to gather results of a smarphone in a
    benchmark
    """

    name: str = 'Benchmark name here'

    # There would be strings with user-friendly names of subtests for string
    # representation of an instance of the class
    subtests: Tuple[str] = ()

    def __init__(self, smartphone: 'Smartphone',
                 percentage_diff: Optional[str] = None) -> None:
        self.smartphone = smartphone
        self.percentage_diff = percentage_diff

    def _get_bench_results(self) -> List[int]:
        """
        Returns a list of integers that represent results of a smartphone in
        subtests of a given benchmark
        """
        return [v for (k, v) in self.get_items() if 'score' in k]

    def get_keys(self):
        return self.__dict__.keys()

    def get_values(self):
        return self.__dict__.values()

    def get_items(self):
        return self.__dict__.items()

    def __getitem__(self, value):
        return getattr(self, value)

    def __setitem__(self, key, value):
        setattr(self, key, value)

    def __str__(self) -> str:
        """
        Returns a string that represents vital information about an instance
        of the class
        """
        score_attrs = self._get_bench_results()

        response = (f'Smartphone {self.smartphone.name}, '
                    f'results in {__class__.name}:\n')
        response += f'Date: {self.smartphone.date}\n'

        # Concatenate a name of the subtest and the score of a smartphone in
        # this subtest
        for name, value in zip(__class__.subtests, score_attrs):
            response += f'{name}: {value}\n'

        return response


@dataclass
class TableReadingSettings:

    """
    Class to store settings for reading a specific table from
    an Excel sheet
    """
    sheet_name: str
    table_start_row: int
    column_with_name: int
    columns_after_name: int
    bench_class: Type[Benchmark]
    bench_attr: str
    chip_or_capacity: str

    def get_values(self):
        return self.__dict__.values()


class GeekBench4(Benchmark):

    """
    Class that represents famous mobile GeekBench 4 benchmark
    """

    name: str = 'GeekBench 4'
    subtests: Tuple[str] = ('Single-Core Score', 'Multi-Core Score',
                            'Total Score')

    def __init__(self, smartphone: 'Smartphone', multi_core_score: int = 0,
                 single_core_score: int = 0) -> None:

        super().__init__(smartphone)

        self.multi_core_score = multi_core_score
        self.single_core_score = single_core_score
        self.total_score: int = multi_core_score + single_core_score


class SlingShotExtreme(Benchmark):
    """
    Class that represents famous mobile 3DMark benchmark, specifically
    Sling Shot Extreme subtest
    """

    name: str = '3DMark Sling Shot Extreme'
    subtests: Tuple[str] = ('Score',)

    def __init__(self, smartphone: 'Smartphone', score: int = 0) -> None:
        super().__init__(smartphone)
        self.total_score = score


class Antutu7(Benchmark):

    """
    Class that represents famous mobile Antutu 7 benchmark
    """

    name: str = 'AnTuTu Benchmark 7'
    subtests: Tuple[str] = ('Score',)

    def __init__(self, smartphone: 'Smartphone', score: int = 0) -> None:
        super().__init__(smartphone)
        self.total_score = score


class BatteryTest(Benchmark):

    """
    Class that represents our own testing of a battery life
    Score are minutes - how much each smarpthone lasted in each activity
    """

    name: str = 'Battery Test'
    subtests: Tuple[str] = ('Read score', 'Movie score', 'Game score',
                            'Total score')

    def __init__(self, smartphone: 'Smartphone', movie_score: int = 0,
                 read_score: int = 0, game_score: int = 0):
        super().__init__(smartphone)
        self.read_score = read_score
        self.movie_score = movie_score
        self.game_score = game_score
        self.total_score: int = movie_score + read_score + game_score


class Smartphone:

    """
    Class that contains info about a smartphone: name, chip, battery capacity
    and results in benchmarks
    """

    def __init__(self, name: str,
                 date: Optional[datetime] = None,
                 highlight: bool = False,
                 ignore: bool = False) -> None:

        self.name = name
        self.date = date
        self.ignore = ignore  # do not include to a plot if True
        self.highlight = highlight  # highlight in a plot if True
        self.chip: Optional[str] = None
        self.battery_capacity: Optional[str] = None
        self.geek_bench4: GeekBench4 = GeekBench4(self)
        self.sling_shot_extreme: SlingShotExtreme = SlingShotExtreme(self)
        self.antutu7: Antutu7 = Antutu7(self)
        self.battery_test = BatteryTest(self)

    def get_keys(self):
        return self.__dict__.keys()

    def get_values(self):
        return self.__dict__.values()

    def get_items(self):
        return self.__dict__.items()

    def __getitem__(self, value):
        return getattr(self, value)

    def __setitem__(self, key, value):
        setattr(self, key, value)

    def __str__(self) -> str:
        return (f'Smartphone {self.name} on {self.chip} with '
                f'{self.battery_capacity} tested on {self.date}')


class Smartphones:

    """
    Special class to manage instances of a Smartphone class

    It creates new Smartphone instances by reading particular Excel file,
    read and fill in scores from benchmarks. Also it can write this data
    to another Excel file - just to be sure it was read right
    """
    def __init__(self) -> None:
        self.all_smartphones: Dict[str, Smartphone] = {}
        self.highlighted_smartphones: List[Smartphone] = []

    @staticmethod
    def _split_name(raw_name: str) -> Tuple[str, str]:
        """
        Method that takes a string and splits its by a particular pattern

        It takes a string like 'Nokia 1 (Snapdragon 450)' or 'Samsung Galaxy
        Note 9 (4000 мАч)" and splits it into two strings - a smartphone name
        and its processor or battery capacity

        """
        regex = re.compile(r'^(.*)\((.*)\)')
        matches = re.search(regex, raw_name).groups()
        name, attr = matches[0].strip(), matches[1].strip()

        return name, attr

    @staticmethod
    def _make_table_reading_settings() -> Dict[str, TableReadingSettings]:

        """
        Method to make different settings for reading different excel sheets
        and tables with benchmark results

        :return: a dictionary with objects of TableReadingSettings class
        """

        geekbench4_trs = TableReadingSettings(sheet_name='GeekBench 4',
                                              table_start_row=4,
                                              column_with_name=32,
                                              columns_after_name=2,
                                              bench_class=GeekBench4,
                                              bench_attr='geek_bench4',
                                              chip_or_capacity='chip')

        sling_shot_extreme_trs = TableReadingSettings(
            sheet_name='3DMark Sling Shot Extreme',
            table_start_row=4,
            column_with_name=29,
            columns_after_name=1,
            bench_class=SlingShotExtreme,
            bench_attr='sling_shot_extreme',
            chip_or_capacity='chip')

        antutu7_trs = TableReadingSettings(sheet_name='Antutu Benchmark 7',
                                           table_start_row=4,
                                           column_with_name=29,
                                           columns_after_name=1,
                                           bench_class=Antutu7,
                                           bench_attr='antutu7',
                                           chip_or_capacity='chip')

        battery_test_trs = TableReadingSettings(sheet_name='battery_test',
                                                table_start_row=4,
                                                column_with_name=34,
                                                columns_after_name=3,
                                                bench_class=BatteryTest,
                                                bench_attr='battery_test',
                                                chip_or_capacity=
                                                'battery_capacity')

        trs = {'geek_bench4': geekbench4_trs,
               'sling_shot_extreme': sling_shot_extreme_trs,
               'antutu7': antutu7_trs,
               'battery_test': battery_test_trs}

        return trs

    def _from_benchmark_table(self, trs: 'TableReadingSettings') -> None:

        """
        Function that read benchmark scores from different sheet of specific
        Excel file

        :param trs: instance of the TableReadingSettings class
        """

        # Get different setting for different benchmarks
        sheet_name, table_start_row, column_with_name, columns_after_name, \
        bench_class, bench_attr, chip_or_capacity = trs.get_values()

        # where my table with results of smartphones begins
        sheet = get_spreadsheet(sheet_name)

        # read the table row by row to read results, make respective classes
        for lst in sheet:

            raw_name = lst[2]

            # split name from an additional characteristic in parentheses
            name, characteristic = self._split_name(raw_name)
            # get a smartphone object from the dict or create new one
            if name not in self.all_smartphones.keys():
                smartphone = Smartphone(name)
                self.all_smartphones[name] = smartphone
                print(f'ADD {name} from {sheet_name}')
            else:
                smartphone = self.all_smartphones[name]
                print(f'UPDATE {name} from {sheet_name}')

            # set chip or battery capacity value to a smartphone object
            smartphone[chip_or_capacity] = characteristic

            # set date if it hasn't been done yet and if it is possible
            if not smartphone.date:
                if lst[1]:
                    smartphone.date = datetime.strptime(lst[1], '%d.%m.%Y')

            action = lst[0]
            if action == '+':
                # highlight the smartphone on plots
                smartphone.highlight = True
                self.highlighted_smartphones.append(smartphone)
            elif action == '-':
                # don't show the smartphone on plots
                smartphone.ignore = True

            # gather results of a benchmark from a row in Excel
            results = lst[3:-1] if len(lst[3:]) > 1 else [lst[3]]
            for i, r in enumerate(results):
                results[i] = int(results[i])
            bench = bench_class(smartphone, *results)
            smartphone[bench_attr] = bench

    def _evaluate_percentage_difference(self, bench, smartphones):
        # TODO make docstring and type hints
        # for a smartphones to be highlighted we set 100% by default
        if len(self.highlighted_smartphones) == 1:
            best_smartphone = self.highlighted_smartphones[0]
            ref_score = best_smartphone[bench]['total_score']

        # evaluating a smartphone with the best score out of scores of
        # highlighted smartphones
        elif len(self.highlighted_smartphones) > 1:
            best_smartphone = self.highlighted_smartphones[0]
            ref_score = best_smartphone[bench]['total_score']

            for smartphone in self.highlighted_smartphones[1:]:
                score = smartphone[bench]['total_score']
                if score > ref_score:
                    ref_score = score
                    best_smartphone = smartphone

        else:
            # if there is not smartphones to be highlighted, we set 100%
            # just to the one with the highest score in bench
            best_smartphone = smartphones[-1]
            ref_score = smartphones[-1][bench]['total_score']

        # set 100% difference for a reference smartphone
        best_smartphone[bench]['percentage_diff'] = '100'

        for s in smartphones:
            score = s[bench]['total_score']
            difference = int(score * 100 / ref_score)
            if score < ref_score:
                percentage_diff = f'-{100 - difference}'
            elif score > ref_score:
                percentage_diff = f'+{difference - 100}'
            else:
                percentage_diff = str(difference)
            s[bench]['percentage_diff'] = percentage_diff

    def prepare_data(self, benchmark: str) -> DataForPlots:

        # make a list of smartphones with date and values of a benchmark
        # of interest
        smartphones = [x for x in self.all_smartphones.values()
                       if x.date and x[benchmark]['total_score']
                       and not x.ignore]

        # Sort smartphones by date and take only the last 30 ones otherwise
        # a plot will be way too big
        smartphones = sorted(smartphones, key=lambda x: x.date)[-30:]

        # sort the smartphones by the total score in benchmarks
        smartphones.sort(key=lambda x: x[benchmark]['total_score'])

        self._evaluate_percentage_difference(benchmark, smartphones)

        data_for_plots = DataForPlots()

        # find indexes of smartphones that you want to highlight
        for i, smrtphn in enumerate(smartphones):
            if smrtphn.highlight:
                data_for_plots.highlighted_smartphones.append(i)

        attr = 'chip' if benchmark != 'battery_test' else 'battery_capacity'

        for indx, s in enumerate(smartphones):
            # return name and chip of a smartphone with it index
            # according to its performance in GeekBench 4
            number = len(smartphones) - indx
            percentage = s[benchmark]['percentage_diff']
            if indx in data_for_plots.highlighted_smartphones:
                if percentage == '100':
                    strng = f'<b>{number}. {s.name} ({s[attr]})</b>'
                else:
                    strng = (f'<b>{number}. {s.name} ({s[attr]}) '
                             f'({percentage}%)</b>')
            else:
                if percentage == '100':
                    strng = f'{number}. {s.name} ({s[attr]})'
                else:
                    strng = (f'{number}. {s.name} ({s[attr]}) '
                             f'({percentage}%)')

            data_for_plots.y_axis_names.append(strng)

            if benchmark == 'geek_bench4':

                mc_score = s[benchmark]['multi_core_score']
                data_for_plots.x_axis_values.append(mc_score)

                sc_score = s[benchmark]['single_core_score']
                data_for_plots.x_axis_values2.append(sc_score)

            elif benchmark == 'sling_shot_extreme' or benchmark == 'antutu7':

                total_score = s[benchmark]['total_score']
                data_for_plots.x_axis_values.append(total_score)

            elif benchmark == 'battery_test':

                data_for_plots.x_axis_values.append(s[benchmark]['read_score'])
                data_for_plots.x_axis_values2.append(
                    s[benchmark]['movie_score'])
                data_for_plots.x_axis_values3.append(
                    s[benchmark]['game_score'])

        return data_for_plots

    def read_from_excel_book(self) -> None:
        """
        Read several Excel sheets one by one
        :return: None
        """
        trs = self._make_table_reading_settings()
        for bench in cfg.LIST_OF_BENCHS:
            self._from_benchmark_table(trs[bench])

    def write_to_excel(self) -> None:

        """
        Save data about smartphones and their benchmark scores to an Excel
        table
        """

        smartphones = list(self.all_smartphones.values())

        # Prepare workbook and a sheet
        wb = Workbook()
        dest_filename = 'smartphones.xlsx'
        ws1 = wb.active
        ws1.title = 'smartphones'

        heading = ["Date", "Smartphone", "Chip", "Battery"]

        # Supplement heading with names of the benchmarks and their subtests
        benchmark_subclasses = Benchmark.__subclasses__()
        for bench in benchmark_subclasses:
            name = getattr(bench, 'name')
            for subtest in getattr(bench, 'subtests'):
                heading.append(f'{name}: {subtest} ')

        # Write the heading to the Excel sheet
        for column, value in zip(range(1, len(heading)+1, 1), heading):
            ws1.cell(row=1, column=column, value=value)

        # Make a row with the data about a smartphone
        smartphone_row = []
        for i, s in enumerate(smartphones):
            smartphone_row.extend([s.date, s.name, s.chip, s.battery_capacity])

            # Get results of benchmarking a smartphone from its class
            for bench_name in cfg.LIST_OF_BENCHS:
                bench = s[bench_name]
                for k, v in bench.__dict__.items():
                    if 'score' in k:
                        smartphone_row.append(v)

            # Write results of benchmarking of a smartphone to the Excel sheet
            column_len = (range(1, len(smartphone_row) + 1, 1))
            for column, value in zip(column_len, smartphone_row):
                ws1.cell(row=i+2, column=column, value=value)
            smartphone_row[:] = []

        wb.save(filename=dest_filename)
        print(f'Wrote down data to {dest_filename}')


def main():
    smartphones = Smartphones()
    smartphones.read_from_excel_book()


if __name__ == '__main__':
    main()
